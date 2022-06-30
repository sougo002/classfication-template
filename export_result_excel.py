import os
from pathlib import Path

from pytorch_lightning.utilities.seed import seed_everything
import hydra
from hydra.core.hydra_config import HydraConfig
from hydra.utils import get_original_cwd
from omegaconf import DictConfig, OmegaConf
import pandas as pd
import openpyxl

from src.utils.custom_logging import CustomLogger

logger = CustomLogger(__name__).get_logger()


@hydra.main(config_path='config')
def export_as_excel(cfg: DictConfig):
    config_name = HydraConfig.get().job.config_name
    seed_everything(cfg.General.seed)
    fold = cfg.Data.dataset.fold

    # 定数設定
    logger.info('-------------in export excel------------')
    logger.info(f'config - \n{OmegaConf.to_yaml(cfg)}')
    root_dir = Path(get_original_cwd())
    experiment_dir = Path('./')
    logger.info(f'root dir : {root_dir}')
    logger.info(f'experiment dir : {experiment_dir.resolve()}')
    logger.info(f'config : {config_name}')
    result = pd.read_csv(root_dir / cfg.Evaluate.result_file)
    save_dir = root_dir / 'outputs' / config_name
    classes = list(cfg.Model.classes)
    normal_names = [classes[0]]
    SHEET_TITLE = '検証結果'  # シート名の設定
    RESULT_FILE_NAME = f'簡易AI検証結果_{config_name}.xlsx'  # 結果を保存するファイル名

    for index, class_name in enumerate(classes):
        result.loc[result['label'] == index, 'label'] = class_name
        result.loc[result['prediction'] == index, 'prediction'] = class_name

    # ワークブック設定
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title=SHEET_TITLE)  # シートを追加
    ws = wb[SHEET_TITLE]  # 追加したシートを編集対象にする

    # セルの列幅を変更
    ws.column_dimensions['A'].width = 20  # px じゃない
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    for index, name in enumerate(['元画像', 'ヒートマップ', '正解ラベル', 'AI分類結果', '画像名'] + classes):
        ws.cell(row=23, column=index+1).value = name

    for idx, row in enumerate(result.itertuples()):
        label = row[1]
        prediction = row[2]
        image_path = row[3]
        scores = row[4:-1]
        heatmap_path = row[-1]

        idx = idx + 24

        # セルの行幅を変更
        ws.row_dimensions[idx].height = 150  # px
        # 画像を挿入
        img = openpyxl.drawing.image.Image(root_dir / image_path.replace('\\', os.sep))
        img.height = 150
        img.width = 150
        terget_cell = 'A' + str(idx)
        ws.add_image(img, terget_cell)  # シートに画像貼り付け

        heatmap_img = openpyxl.drawing.image.Image(root_dir / heatmap_path.replace('\\', os.sep))
        heatmap_img.height = 150
        heatmap_img.width = 150
        terget_cell = 'B' + str(idx)
        ws.add_image(heatmap_img, terget_cell)  # シートに画像貼り付け

        ws.cell(row=idx, column=3).value = label

        ws.cell(row=idx, column=4).value = prediction

        if label in normal_names and prediction not in normal_names:
            # 過検出
            ws.cell(row=idx, column=4).fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='FCE3EA', bgColor='FCE3EA')
        elif label not in normal_names and prediction in normal_names:
            # 見逃し
            ws.cell(row=idx, column=4).fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='DCE6F1', bgColor='DCE6F1')

        ws.cell(row=idx, column=5).value = Path(image_path).name
        for i, score in enumerate(scores):
            ws.cell(row=idx, column=6+i).value = score
    wb.save(save_dir / RESULT_FILE_NAME)


if __name__ == '__main__':
    export_as_excel()
